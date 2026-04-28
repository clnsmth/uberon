"""
Stage 1: Generate initial ROBOT template TSVs from HRA ASCTB unmapped terms.

Each input row is pre-classified as a leaf or group term (linguistic rules) and routed
to the appropriate template:
  - Leaf terms → standard template with SC (asserted is_a/part_of)
  - Group terms → groups template with EC (equivalent class: genus + part_of some Y)

Input:  An xlsx file (default: hra_unmapped-asct-term-list-with-refs.xlsx at repo root)
        OR a pre-exported CSV with the same columns as the 'as-temp terms' sheet.
        Optionally filter to a specific ASCTB table (e.g. 'muscular-system').

Outputs  (REPO_ROOT = two levels up from this script):
  bulk_ntr_workflow/outputs/template_initial.tsv         — leaf working template
  bulk_ntr_workflow/outputs/template_groups_initial.tsv  — groups working template
  src/templates/<name>.template.tsv                      — leaf final template
  src/templates/<name>-groups.template.tsv               — groups final template
  src/templates/<name>-reports/input.tsv                 — filtered input rows + term_type
  src/templates/<name>-reports/errors.tsv                — input problems
  src/templates/<name>-reports/candidates.tsv            — pre-mapped existing terms

Usage:
  cd bulk_ntr_workflow
  uv run --with openpyxl scripts/generate_template.py \\
    --input ../hra_unmapped-asct-term-list-with-refs.xlsx \\
    --table muscular-system \\
    --name hra-muscular \\
    --start-id 9900001
"""

import argparse
import csv
import re
import sys
from datetime import date
from pathlib import Path

# bulk_ntr_workflow/scripts/ → bulk_ntr_workflow/ → repo root
NTR_ROOT  = Path(__file__).resolve().parent.parent
REPO_ROOT = NTR_ROOT.parent

WORK_DIR        = NTR_ROOT / "outputs"
WORK_DIR.mkdir(parents=True, exist_ok=True)
WORK_TSV        = WORK_DIR / "template_initial.tsv"             # default leaf
WORK_GROUPS_TSV = WORK_DIR / "template_groups_initial.tsv"       # groups
# System overlay working files: outputs/template_<overlay>_initial.tsv (created on demand)

# ROBOT template column headers and directives — DEFAULT LEAF template (asserted SC)
# Phase 6: develops_from is OPTIONAL — empty cell ⇒ no axiom emitted by ROBOT
TEMPLATE_HEADERS = [
    "ID", "LABEL", "Definition", "def_xref",
    "is_a", "part_of", "develops_from",
    "In_subset", "Date", "Contributor", "Present_in_taxon",
    "Wikipedia_image", "xref",
]
TEMPLATE_DIRECTIVES = [
    "ID", "LABEL", "A IAO:0000115", ">A oboInOwl:hasDbXref SPLIT=|",
    "SC %", "SC BFO:0000050 some %", "SC RO:0002202 some %",
    "AI oboInOwl:inSubset", "AT dcterms:date^^xsd:dateTime",
    "AI dcterms:contributor", "AI RO:0002175",
    "A foaf:depiction", "A oboInOwl:hasDbXref SPLIT=|",
]

# Phase 7: MUSCLE LEAF template overlay — adds muscle-specific relations.
# RO IDs: has_muscle_origin=RO:0002372, has_muscle_insertion=RO:0002373, innervated_by=RO:0002005
# Inserted between develops_from and In_subset (positions 7-9).
MUSCLE_TEMPLATE_HEADERS = TEMPLATE_HEADERS[:7] + [
    "has_muscle_origin", "has_muscle_insertion", "innervated_by",
] + TEMPLATE_HEADERS[7:]
MUSCLE_TEMPLATE_DIRECTIVES = TEMPLATE_DIRECTIVES[:7] + [
    "SC RO:0002372 some %", "SC RO:0002373 some %", "SC RO:0002005 some %",
] + TEMPLATE_DIRECTIVES[7:]

# Map source-table value to a system overlay name. Unmapped tables → 'default'.
# Future overlays for skeleton, vasculature, nervous-system go here (see ROADMAP).
SYSTEM_OVERLAYS = {
    "muscular-system": "muscle",
}

# Per-overlay header/directive sets
OVERLAY_TEMPLATES = {
    "default": (TEMPLATE_HEADERS,        TEMPLATE_DIRECTIVES),
    "muscle":  (MUSCLE_TEMPLATE_HEADERS, MUSCLE_TEMPLATE_DIRECTIVES),
}


def classify_system(record: dict) -> str:
    """Return the system overlay name for a row; 'default' if no overlay applies."""
    return SYSTEM_OVERLAYS.get(record.get("table", ""), "default")


def overlay_paths(overlay: str, name: str) -> tuple[Path, Path]:
    """Return (working_tsv, final_tsv) paths for a given overlay name."""
    templates_dir = REPO_ROOT / "src" / "templates"
    if overlay == "default":
        work  = WORK_DIR / "template_initial.tsv"
        final = templates_dir / f"{name}.template.tsv"
    else:
        work  = WORK_DIR / f"template_{overlay}_initial.tsv"
        final = templates_dir / f"{name}-{overlay}.template.tsv"
    return work, final


# ROBOT template — GROUPS template (equivalent class: genus + part_of some Y)
GROUPS_TEMPLATE_HEADERS = [
    "ID", "LABEL", "Definition", "def_xref",
    "genus", "location",
    "In_subset", "Date", "Contributor", "Present_in_taxon",
    "Wikipedia_image", "xref",
]
GROUPS_TEMPLATE_DIRECTIVES = [
    "ID", "LABEL", "A IAO:0000115", ">A oboInOwl:hasDbXref SPLIT=|",
    "EC %", "EC BFO:0000050 some %",
    "AI oboInOwl:inSubset", "AT dcterms:date^^xsd:dateTime",
    "AI dcterms:contributor", "AI RO:0002175",
    "A foaf:depiction", "A oboInOwl:hasDbXref SPLIT=|",
]

# Columns for input.tsv (mirrors the raw source columns + term_type pre-classification)
INPUT_HEADERS = [
    "table", "as_iri", "label", "uberon_id",
    "parent_id", "parent_label", "references", "term_type",
]

# Columns for errors.tsv
ERROR_HEADERS = ["label", "as_iri", "issue_type", "parent_id", "parent_label", "detail"]

# Columns for candidates.tsv
CANDIDATE_HEADERS = ["label", "as_iri", "uberon_id", "note"]

SUBSET_IRI    = "http://purl.obolibrary.org/obo/uberon/core#added_by_HRA"
CREATION_DATE = f"{date.today().isoformat()}T00:00:00Z"
TAXON_IRI     = "http://purl.obolibrary.org/obo/NCBITaxon_9606"

ORCID_RE = re.compile(r'^https://orcid\.org/\d{4}-\d{4}-\d{4}-\d{3}[\dX]$')

DEFAULT_START_ID = 9900001

UBERON_RE  = re.compile(r'^UBERON:\d{7}$')
FMA_IRI_RE = re.compile(r'fma/fma(\d+)', re.IGNORECASE)

# Linguistic patterns for grouping terms (collective classes, not specific named entities).
# When matched, the term is routed to the groups template (EquivalentClass form).
# Default if none match: "leaf" (asserted SC subclass).
GROUP_PATTERNS = [
    re.compile(r'\bmuscle of (?!the )', re.IGNORECASE),
    re.compile(r'\b(pelvic floor|thoracic wall|abdominal wall|chest|chest wall) muscle\b', re.IGNORECASE),
    re.compile(r'\b(dorsum|sole) of (foot|hand) muscle\b', re.IGNORECASE),
    re.compile(r'\b(circular|longitudinal) pharyngeal muscle\b', re.IGNORECASE),
    re.compile(r'\b(intrinsic|extrinsic) (eye|ear|tongue|hand|foot|laryngeal|lingual) muscle\b', re.IGNORECASE),
    re.compile(r'\b(hypothenar|thenar) hand muscle\b', re.IGNORECASE),
    re.compile(r'\b(palmar|plantar) interosseous muscle\b', re.IGNORECASE),
    re.compile(r'\b(superficial|intermediate|deep) back muscle\b', re.IGNORECASE),
    re.compile(r'\b(anterior|posterior|lateral|medial) vertebral muscle\b', re.IGNORECASE),
    re.compile(r'\b(anterior|posterior|lateral|medial) compartment( of \w+)? muscle\b', re.IGNORECASE),
    re.compile(r'\b(spinotransversales|segmental back|external ear|middle ear|cranial) muscle\b', re.IGNORECASE),
    re.compile(r'\b(posterior|anterior|lateral|medial) abdominal wall muscle\b', re.IGNORECASE),
    re.compile(r'\bmuscle of (facial expression|mastication)\b', re.IGNORECASE),
]

# Subdivision patterns — head/belly/part/portion/crus/etc of a named muscle → leaf
LEAF_PART_PATTERNS = [
    re.compile(r'\b(head|belly|part|portion|crus|fascicle|layer|zone|lamina) of\b', re.IGNORECASE),
]


def classify_term_type(label: str) -> str:
    """Classify a term label as 'group' or 'leaf' using linguistic rules.

    Default: 'leaf'. A term matching any LEAF_PART_PATTERN (e.g. 'X head of Y muscle')
    is always 'leaf', even if a GROUP_PATTERN would otherwise match. Specific
    subdivisions of a named structure trump grouping cues.
    """
    if not label:
        return "leaf"
    for pat in LEAF_PART_PATTERNS:
        if pat.search(label):
            return "leaf"
    for pat in GROUP_PATTERNS:
        if pat.search(label):
            return "group"
    return "leaf"


# ---------------------------------------------------------------------------
# Input reading
# ---------------------------------------------------------------------------

def read_xlsx(path: Path, table_filter: str | None) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl not installed — run: uv run --with openpyxl ...")

    wb = openpyxl.load_workbook(str(path), read_only=True)
    if "as-temp terms" not in wb.sheetnames:
        sys.exit(f"Sheet 'as-temp terms' not found in {path}. Sheets: {wb.sheetnames}")
    ws = wb["as-temp terms"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("Sheet is empty")

    raw_headers = [str(h).strip() if h else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(raw_headers)}

    def get(row, name, default=""):
        idx = col.get(name)
        if idx is None:
            return default
        v = row[idx]
        return str(v).strip() if v is not None else default

    records = []
    for row in rows[1:]:
        if not any(row):
            continue
        table = get(row, "tables")
        if table_filter and table != table_filter:
            continue
        records.append({
            "table":        table,
            "iri":          get(row, "as"),
            "label":        get(row, "as_label"),
            "uberon_id":    get(row, "UBERON ID"),
            "parent_id":    get(row, "parents_as"),
            "parent_label": get(row, "parents_as_label"),
            "references":   get(row, "references"),
        })
    return records


def read_csv(path: Path, table_filter: str | None) -> list[dict]:
    records = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            table = row.get("tables", "").strip()
            if table_filter and table != table_filter:
                continue
            records.append({
                "table":        table,
                "iri":          row.get("as", "").strip(),
                "label":        row.get("as_label", "").strip(),
                "uberon_id":    row.get("UBERON ID", "").strip(),
                "parent_id":    row.get("parents_as", "").strip(),
                "parent_label": row.get("parents_as_label", "").strip(),
                "references":   row.get("references", "").strip(),
            })
    return records


# ---------------------------------------------------------------------------
# Parent ID classification
# ---------------------------------------------------------------------------

def classify_parent(parent_id: str) -> str:
    """Return 'uberon', 'fma', 'asctb_temp', or 'unknown'."""
    if UBERON_RE.match(parent_id):
        return "uberon"
    if FMA_IRI_RE.search(parent_id):
        return "fma"
    if "ASCTB-TEMP" in parent_id or "asctb-temp" in parent_id.lower():
        return "asctb_temp"
    return "unknown"


def fma_id_from_iri(iri: str) -> str:
    m = FMA_IRI_RE.search(iri)
    return f"FMA:{m.group(1)}" if m else iri


# ---------------------------------------------------------------------------
# Reference formatting  (comma-separated → pipe-separated)
# ---------------------------------------------------------------------------

def format_refs(raw: str, asctb_iri: str) -> str:
    parts = [r.strip() for r in raw.split(",") if r.strip()]
    if asctb_iri and asctb_iri not in parts:
        parts.append(asctb_iri)
    return "|".join(parts) if parts else ""


# ---------------------------------------------------------------------------
# TSV helpers
# ---------------------------------------------------------------------------

def write_tsv(path: Path, headers: list[str], rows: list[list]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter="\t")
        writer.writerow(headers)
        writer.writerows(rows)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def resolve_contributor(contributor_arg: str | None) -> str:
    """Return a validated ORCID IRI, prompting if not supplied."""
    if contributor_arg:
        iri = contributor_arg.strip()
        if not iri.startswith("https://orcid.org/"):
            iri = f"https://orcid.org/{iri}"
        if not ORCID_RE.match(iri):
            sys.exit(f"Invalid ORCID format: {iri}\nExpected: https://orcid.org/XXXX-XXXX-XXXX-XXXX")
        return iri
    while True:
        raw = input("Contributor ORCID (e.g. https://orcid.org/0000-0000-0000-0000): ").strip()
        if not raw.startswith("https://orcid.org/"):
            raw = f"https://orcid.org/{raw}"
        if ORCID_RE.match(raw):
            return raw
        print(f"  Invalid format, try again.")


def process(input_path: Path, table_filter: str | None, start_id: int, name: str,
            contributor_iri: str, limit: int | None = None) -> None:
    suffix = input_path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        records = read_xlsx(input_path, table_filter)
    elif suffix in (".csv", ".tsv"):
        records = read_csv(input_path, table_filter)
    else:
        sys.exit(f"Unsupported file type: {suffix}")

    if not records:
        sys.exit("No records found (check --table filter)")

    if limit is not None:
        records = records[:limit]

    # Output paths
    templates_dir    = REPO_ROOT / "src" / "templates"
    reports_dir      = templates_dir / f"{name}-reports"
    final_groups_tsv = templates_dir / f"{name}-groups.template.tsv"
    input_tsv        = reports_dir / "input.tsv"
    errors_tsv       = reports_dir / "errors.tsv"
    candidates_tsv   = reports_dir / "candidates.tsv"
    reports_dir.mkdir(parents=True, exist_ok=True)

    # Step 0: rows are partitioned by system overlay (default vs muscle vs ...).
    # leaf_rows_by_overlay[overlay] holds the rows destined for that overlay's template.
    leaf_rows_by_overlay: dict[str, list] = {}
    group_rows    = []
    error_rows    = []
    candidate_rows = []
    input_rows    = []
    counter       = start_id

    for rec in records:
        label      = rec["label"]
        iri        = rec["iri"]
        uberon_id  = rec["uberon_id"]
        parent_id  = rec["parent_id"]
        parent_lbl = rec["parent_label"]
        refs       = rec["references"]

        term_type = classify_term_type(label) if label else "leaf"

        # Save to input.tsv regardless of outcome (now includes term_type)
        input_rows.append([rec["table"], iri, label, uberon_id,
                           parent_id, parent_lbl, refs, term_type])

        if not label:
            error_rows.append(["", iri, "missing_label", "", "", ""])
            continue

        # Already mapped — skip from template, log as candidate
        if uberon_id and UBERON_RE.match(uberon_id):
            candidate_rows.append([label, iri, uberon_id, "pre-assigned in input"])
            continue

        # Classify parent
        parent_class = classify_parent(parent_id) if parent_id else "unknown"

        if parent_class == "uberon":
            # Embed parent ID; subagent resolves is_a vs part_of
            is_a_val    = f"INFER:{parent_id}"
            part_of_val = f"INFER:{parent_id}"
        elif parent_class == "fma":
            fma_curie   = fma_id_from_iri(parent_id)
            is_a_val    = f"NEEDS_MAPPING:{fma_curie}"
            part_of_val = f"NEEDS_MAPPING:{fma_curie}"
            error_rows.append([
                label, iri, "fma_parent",
                fma_curie, parent_lbl,
                "Subagent should resolve FMA→UBERON via OLS4"
            ])
        elif parent_class == "asctb_temp":
            # Embed parent label so subagent can try OLS4 to find correct UBERON parent
            safe_lbl    = parent_lbl.replace("|", ";")
            is_a_val    = f"UNRESOLVABLE:{safe_lbl}"
            part_of_val = f"UNRESOLVABLE:{safe_lbl}"
            error_rows.append([
                label, iri, "asctb_temp_parent",
                parent_id, parent_lbl,
                "Parent not yet in UBERON; subagent should search OLS4 for correct parent"
            ])
        else:
            is_a_val    = "UNKNOWN"
            part_of_val = "UNKNOWN"
            error_rows.append([
                label, iri, "unknown_parent",
                parent_id, parent_lbl, "Unrecognised parent ID format"
            ])

        def_xref = format_refs(refs, iri)
        # Pre-populate xref with FMA ID if the term's own IRI is an FMA IRI
        own_fma = fma_id_from_iri(iri) if FMA_IRI_RE.search(iri) else ""

        if term_type == "group":
            # Groups template: genus + location columns are populated by the subagent
            group_rows.append([
                f"http://purl.obolibrary.org/obo/UBERON_{counter}",
                label,
                "[PENDING]",
                def_xref,
                "",       # genus — filled by subagent
                "",       # location — filled by subagent
                SUBSET_IRI,
                CREATION_DATE,
                contributor_iri,
                TAXON_IRI,
                "",       # Wikipedia_image — filled by subagent
                own_fma,  # xref — FMA from source IRI; subagent appends
            ])
        else:
            overlay = classify_system(rec)
            base_row = [
                f"http://purl.obolibrary.org/obo/UBERON_{counter}",
                label,
                "[PENDING]",
                def_xref,
                is_a_val,
                part_of_val,
                "",       # develops_from — filled by subagent if applicable
            ]
            if overlay == "muscle":
                base_row += ["", "", ""]  # has_muscle_origin, has_muscle_insertion, innervated_by
            base_row += [
                SUBSET_IRI,
                CREATION_DATE,
                contributor_iri,
                TAXON_IRI,
                "",       # Wikipedia_image — filled by subagent
                own_fma,
            ]
            leaf_rows_by_overlay.setdefault(overlay, []).append(base_row)
        counter += 1

    # Write per-overlay LEAF working + final templates
    overlay_summary = []
    for overlay, rows in sorted(leaf_rows_by_overlay.items()):
        headers, directives = OVERLAY_TEMPLATES[overlay]
        work_path, final_path = overlay_paths(overlay, name)
        for path in (work_path, final_path):
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f, delimiter="\t")
                writer.writerow(headers)
                writer.writerow(directives)
                writer.writerows(rows)
        overlay_summary.append((overlay, len(rows), final_path))

    # Write GROUPS working + final templates
    for path in (WORK_GROUPS_TSV, final_groups_tsv):
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter="\t")
            writer.writerow(GROUPS_TEMPLATE_HEADERS)
            writer.writerow(GROUPS_TEMPLATE_DIRECTIVES)
            writer.writerows(group_rows)

    # Write reports
    write_tsv(input_tsv, INPUT_HEADERS, input_rows)
    write_tsv(errors_tsv, ERROR_HEADERS, error_rows)
    write_tsv(candidates_tsv, CANDIDATE_HEADERS, candidate_rows)

    # Routing summary (Step 0)
    parts = ", ".join(f"{ov}={n}" for ov, n, _ in overlay_summary) or "(none)"
    print(f"Step 0 routing: {parts}, group={len(group_rows)}")
    print()
    for overlay, n, final_path in overlay_summary:
        work_path, _ = overlay_paths(overlay, name)
        print(f"Leaf template [{overlay}] → {final_path}  ({n} rows)")
    print(f"Groups template → {final_groups_tsv}  ({len(group_rows)} rows)")
    print(f"Reports         → {reports_dir}/")
    print(f"  input.tsv      {len(input_rows)} rows")
    print(f"  errors.tsv     {len(error_rows)} rows")
    print(f"  candidates.tsv {len(candidate_rows)} rows")

    total_leaf = sum(len(r) for r in leaf_rows_by_overlay.values())
    uberon_p = sum(1 for r in records if classify_parent(r["parent_id"]) == "uberon")
    fma_p    = sum(1 for r in records if classify_parent(r["parent_id"]) == "fma")
    asctb_p  = sum(1 for r in records if classify_parent(r["parent_id"]) == "asctb_temp")
    print(f"\nTemplate rows: leaf={total_leaf} group={len(group_rows)} | "
          f"Parents: UBERON={uberon_p} FMA={fma_p} ASCTB-TEMP={asctb_p}")
    if asctb_p:
        print(f"  ⚠ {asctb_p} terms have ASCTB-TEMP parents — "
              f"subagent will attempt OLS4 lookup for correct parent")


def main():
    parser = argparse.ArgumentParser(
        description="Generate initial UBERON NTR ROBOT template from HRA ASCTB unmapped terms"
    )
    parser.add_argument(
        "--input", required=True,
        help="Path to input xlsx or csv file"
    )
    parser.add_argument(
        "--table", default=None,
        help="Filter to a specific ASCTB table name (e.g. 'muscular-system')"
    )
    parser.add_argument(
        "--name", required=True,
        help="Template name used for output filenames (e.g. 'hra-muscular')"
    )
    parser.add_argument(
        "--start-id", type=int, default=DEFAULT_START_ID,
        help=f"Starting UBERON:99xxxxx counter (default: {DEFAULT_START_ID})"
    )
    parser.add_argument(
        "--limit", type=int, default=None,
        help="Process only the first N terms (for testing)"
    )
    parser.add_argument(
        "--contributor", default=None,
        help="Contributor ORCID IRI (e.g. https://orcid.org/0000-0000-0000-0000). "
             "Prompted interactively if omitted."
    )
    args = parser.parse_args()
    contributor_iri = resolve_contributor(args.contributor)
    process(Path(args.input), args.table, args.start_id, args.name, contributor_iri, args.limit)


if __name__ == "__main__":
    main()
